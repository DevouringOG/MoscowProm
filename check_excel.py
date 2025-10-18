"""Check Excel file structure."""
from pathlib import Path
import openpyxl

excel_file = Path("Тип данных от организаций.xlsx")

if not excel_file.exists():
    print(f"❌ Excel файл не найден: {excel_file}")
    exit(1)

print(f"✅ Excel файл найден: {excel_file}")

workbook = openpyxl.load_workbook(excel_file, data_only=True)
sheet = workbook.active

print(f"\n📊 Информация о листе:")
print(f"  • Название листа: {sheet.title}")
print(f"  • Максимальная строка: {sheet.max_row}")
print(f"  • Максимальная колонка: {sheet.max_column}")

# Check first few rows
print(f"\n📝 Первые 5 строк:")
for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    print(f"\nСтрока {row_idx}: {len(row)} колонок")
    if row_idx == 1:
        # Print first 10 columns of header
        print(f"  Первые 10 колонок:")
        for i, val in enumerate(row[:10]):
            print(f"    [{i}] {val}")
    else:
        # Print key columns for data rows
        print(f"  [0] № = {row[0] if len(row) > 0 else 'НЕТ'}")
        print(f"  [1] ИНН = {row[1] if len(row) > 1 else 'НЕТ'}")
        print(f"  [2] Название = {row[2] if len(row) > 2 else 'НЕТ'}")
        print(f"  [47] Выручка 2017 = {row[47] if len(row) > 47 else 'НЕТ'}")
