"""Service for exporting organizations data to Excel."""
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from app.db.models import Organization, OrganizationMetrics, OrganizationTaxes, OrganizationAssets, OrganizationProducts, OrganizationMeta


def export_organizations_to_excel(organizations: List[Organization], db: Session) -> BytesIO:
    """
    Export list of organizations to Excel file with all columns.
    
    Args:
        organizations: List of Organization objects to export
        db: Database session
        
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Предприятия"
    
    # Define headers
    headers = [
        "№",
        "ИНН",
        "Наименование организации",
        "Полное наименование организации",
        "Статус СПАРК",
        "Статус внутренний",
        "Статус ИТОГ",
        "Дата добавления в реестр",
        "Юридический адрес",
        "Адрес производства",
        "Адрес дополнительной площадки",
        "Основная отрасль",
        "Подотрасль (Основная)",
        "Дополнительная отрасль",
        "Подотрасль (Дополнительная)",
        "Отраслевые презентации",
        "Основной ОКВЭД (СПАРК)",
        "Вид деятельности по основному ОКВЭД (СПАРК)",
        "Производственный ОКВЭД",
        "Вид деятельности по производственному ОКВЭД",
        "Общие сведения об организации",
        "Размер предприятия (итог)",
        "Размер предприятия (итог) 2022",
        "Размер предприятия (по численности)",
        "Размер предприятия (по численности) 2022",
        "Размер предприятия (по выручке)",
        "Размер предприятия (по выручке) 2022",
        "Дата регистрации",
        "Руководитель",
        "Головная организация",
        "ИНН головной организации",
        "Вид отношения головной организации",
        "Контактные данные руководства",
        "Почта руководства",
        "Контакт сотрудника организации",
        "Номер телефона",
        "Контактные данные ответственного по ЧС",
        "Сайт",
        "Электронная почта",
        "Данные о мерах поддержки",
        "Наличие особого статуса",
        "Площадка итог",
        "Получена поддержка от г. Москвы",
        "Системообразующее предприятие",
        "Статус МСП",
        "То самое",
        "Финансово-экономические показатели",
    ]
    
    # Add financial metrics headers for years 2017-2023
    years = list(range(2017, 2024))
    
    for year in years:
        headers.append(f"Выручка предприятия, тыс. руб. {year}")
    
    for year in years:
        headers.append(f"Чистая прибыль (убыток),тыс. руб. {year}")
    
    for year in years:
        headers.append(f"Среднесписочная численность персонала (всего по компании), чел {year}")
    
    for year in years:
        headers.append(f"Среднесписочная численность персонала, работающего в Москве, чел {year}")
    
    for year in years:
        headers.append(f"Фонд оплаты труда всех сотрудников организации, тыс. руб {year}")
    
    for year in years:
        headers.append(f"Фонд оплаты труда сотрудников, работающих в Москве, тыс. руб {year}")
    
    for year in years:
        headers.append(f"Средняя з.п. всех сотрудников организации, тыс.руб. {year}")
    
    for year in years:
        headers.append(f"Средняя з.п. сотрудников, работающих в Москве, тыс.руб. {year}")
    
    # Tax headers for years 2017-2024
    tax_years = list(range(2017, 2025))
    
    for year in tax_years:
        headers.append(f"Налоги, уплаченные в бюджет Москвы (без акцизов), тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"Налог на прибыль, тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"Налог на имущество, тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"Налог на землю, тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"НДФЛ, тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"Транспортный налог, тыс.руб. {year}")
    
    for year in tax_years:
        headers.append(f"Прочие налоги {year}")
    
    for year in tax_years:
        headers.append(f"Акцизы, тыс. руб. {year}")
    
    # Investments and export
    headers.extend([
        "Инвестиции в Мск 2021 тыс. руб.",
        "Инвестиции в Мск 2022 тыс. руб.",
        "Инвестиции в Мск 2023 тыс. руб.",
    ])
    
    export_years = list(range(2019, 2024))
    for year in export_years:
        headers.append(f"Объем экспорта, тыс. руб. {year}")
    
    # Property and assets
    headers.extend([
        "Имущественно-земельный комплекс",
        "Кадастровый номер ЗУ",
        "Площадь ЗУ",
        "Вид разрешенного использования ЗУ",
        "Вид собственности ЗУ",
        "Собственник ЗУ",
        "Кадастровый номер ОКСа",
        "Площадь ОКСов",
        "Вид разрешенного использования ОКСов",
        "Тип строения и цель использования",
        "Вид собственности ОКСов",
        "СобственникОКСов",
        "Площадь производственных помещений, кв.м.",
    ])
    
    # Products
    headers.extend([
        "Производимая продукция",
        "Стандартизированная продукция",
        "Название (виды производимой продукции)",
        "Перечень производимой продукции по кодам ОКПД 2",
        "Перечень производимой продукции по типам и сегментам",
        "Каталог продукции",
        "Наличие госзаказа",
        "Уровень загрузки производственных мощностей",
        "Наличие поставок продукции на экспорт",
        "Объем экспорта (млн.руб.) за предыдущий календарный год",
        "Перечень государств куда экспортируется продукция",
        "Код ТН ВЭД ЕАЭС",
    ])
    
    # Additional
    headers.extend([
        "Развитие Реестра",
        "Отрасль промышленности по Спарк и Справочнику",
        "Координаты юридического адреса",
        "Координаты адреса производства",
        "Координаты адреса дополнительной площадки",
        "Координаты (широта)",
        "Координаты (долгота)",
        "Округ",
        "Район",
    ])
    
    # Write headers
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 40
    
    # Write data
    for idx, org in enumerate(organizations, start=1):
        # Get related data
        metrics_dict = {}
        for metric in db.query(OrganizationMetrics).filter(OrganizationMetrics.organization_id == org.id).all():
            metrics_dict[metric.year] = metric
        
        taxes_dict = {}
        for tax in db.query(OrganizationTaxes).filter(OrganizationTaxes.organization_id == org.id).all():
            taxes_dict[tax.year] = tax
        
        assets = db.query(OrganizationAssets).filter(OrganizationAssets.organization_id == org.id).first()
        products = db.query(OrganizationProducts).filter(OrganizationProducts.organization_id == org.id).first()
        meta = db.query(OrganizationMeta).filter(OrganizationMeta.organization_id == org.id).first()
        
        # Build row data
        row = [
            idx,  # №
            org.inn or "",
            org.name or "",
            org.full_name or "",
            org.status_spark or "",
            org.status_internal or "",
            org.status_final or "",
            org.date_added.strftime("%d.%m.%Y") if org.date_added else "",
            org.legal_address or "",
            org.production_address or "",
            org.additional_address or "",
            org.main_industry or "",
            org.main_subindustry or "",
            org.extra_industry or "",
            org.extra_subindustry or "",
            meta.presentation_links if meta else "",  # Отраслевые презентации
            org.main_okved or "",
            org.main_okved_name or "",
            org.prod_okved or "",
            org.prod_okved_name or "",
            org.company_info or "",
            org.company_size or "",
            org.company_size_2022 or "",
            org.size_by_employees or "",
            org.size_by_employees_2022 or "",
            org.size_by_revenue or "",
            org.size_by_revenue_2022 or "",
            org.registration_date.strftime("%d.%m.%Y") if org.registration_date else "",
            org.head_name or "",
            org.parent_org_name or "",
            org.parent_org_inn or "",
            org.parent_relation_type or "",
            org.head_contacts or "",
            org.head_email or "",
            org.employee_contact or "",
            org.phone or "",
            org.emergency_contact or "",
            org.website or "",
            org.email or "",
            org.support_data or "",
            org.special_status or "",
            org.site_final or "",
            "Да" if org.got_moscow_support else "Нет",
            "Да" if org.is_system_critical else "Нет",
            org.msp_status or "",
            "",  # То самое
            "",  # Финансово-экономические показатели (заголовок)
        ]
        
        # Add financial metrics by year
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.revenue if metric and metric.revenue else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.profit if metric and metric.profit else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.total_employees if metric and metric.total_employees else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.moscow_employees if metric and metric.moscow_employees else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.total_fot if metric and metric.total_fot else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.moscow_fot if metric and metric.moscow_fot else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.avg_salary_total if metric and metric.avg_salary_total else "")
        
        for year in years:
            metric = metrics_dict.get(year)
            row.append(metric.avg_salary_moscow if metric and metric.avg_salary_moscow else "")
        
        # Add tax data by year
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.total_taxes_moscow if tax and tax.total_taxes_moscow else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.profit_tax if tax and tax.profit_tax else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.property_tax if tax and tax.property_tax else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.land_tax if tax and tax.land_tax else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.ndfl if tax and tax.ndfl else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.transport_tax if tax and tax.transport_tax else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.other_taxes if tax and tax.other_taxes else "")
        
        for year in tax_years:
            tax = taxes_dict.get(year)
            row.append(tax.excise if tax and tax.excise else "")
        
        # Investments
        inv_2021 = metrics_dict.get(2021)
        inv_2022 = metrics_dict.get(2022)
        inv_2023 = metrics_dict.get(2023)
        row.extend([
            inv_2021.investments if inv_2021 and inv_2021.investments else "",
            inv_2022.investments if inv_2022 and inv_2022.investments else "",
            inv_2023.investments if inv_2023 and inv_2023.investments else "",
        ])
        
        # Export volumes
        for year in export_years:
            metric = metrics_dict.get(year)
            row.append(metric.export_volume if metric and metric.export_volume else "")
        
        # Property and assets
        row.extend([
            assets.property_summary if assets else "",
            assets.cadastral_number_land if assets else "",
            assets.land_area if assets and assets.land_area else "",
            assets.land_usage if assets else "",
            assets.land_ownership_type if assets else "",
            assets.land_owner if assets else "",
            assets.cadastral_number_building if assets else "",
            assets.building_area if assets and assets.building_area else "",
            assets.building_usage if assets else "",
            assets.building_type if assets else "",
            assets.building_ownership_type if assets else "",
            assets.building_owner if assets else "",
            assets.production_area if assets and assets.production_area else "",
        ])
        
        # Products
        row.extend([
            "",  # Производимая продукция (общий заголовок)
            products.standardized_product if products else "",
            products.product_name if products else "",
            products.okpd2_codes if products else "",
            products.product_types if products else "",
            products.product_catalog if products else "",
            "Да" if products and products.has_government_orders else "Нет",
            products.capacity_usage if products and products.capacity_usage else "",
            "Да" if products and products.has_export else "Нет",
            products.export_volume_last_year if products and products.export_volume_last_year else "",
            products.export_countries if products else "",
            products.tnved_code if products else "",
        ])
        
        # Additional
        row.extend([
            meta.registry_development if meta else "",
            f"{meta.industry_spark or ''} / {meta.industry_directory or ''}" if meta else "",
            org.legal_address_coords or "",
            org.production_address_coords or "",
            org.additional_address_coords or "",
            org.coordinates_lat if org.coordinates_lat else "",
            org.coordinates_lon if org.coordinates_lon else "",
            org.district or "",
            org.region or "",
        ])
        
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
